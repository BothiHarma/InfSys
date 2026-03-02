# scheduling_app/export/excel_exporter.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import io


def create_excel_report(
    summary_data: pd.DataFrame,
    idle_by_machine: pd.DataFrame,
    gantt_before: pd.DataFrame,
    gantt_after: pd.DataFrame,
    completion_by_job: pd.DataFrame,
    processing_matrix: list,  
    pie_idle_png: bytes = None,
    pie_comp_png: bytes = None
) -> bytes:
    output = io.BytesIO()
    wb = Workbook()

    # исходная матрица обработки
    ws_matrix = wb.active
    ws_matrix.title = "Исходная матрица"
    
    # создаём DataFrame из матрицы
    n_jobs = len(processing_matrix)
    n_machines = len(processing_matrix[0]) if processing_matrix else 0
    
    matrix_df = pd.DataFrame(
        processing_matrix,
        index=[f"Деталь {i}" for i in range(n_jobs)],
        columns=[f"Станок {j}" for j in range(n_machines)]
    )
    
    for r in dataframe_to_rows(matrix_df, index=True, header=True):
        ws_matrix.append(r)
    
    # автоширина колонок
    for col in ws_matrix.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_matrix.column_dimensions[column].width = min(adjusted_width, 15)

    # сводка
    ws_summary = wb.create_sheet("Сводка")
    for r in dataframe_to_rows(summary_data, index=False, header=True):
        ws_summary.append(r)
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20

    # простои 
    ws_idle = wb.create_sheet("Простои по станкам")
    idle_df = idle_by_machine.rename(columns={
        "Machine": "Станок",
        "TotalWorkTime": "Время работы",
        "IdleTime": "Простой",
        "Utilization": "Загрузка (%)"
    })
    for r in dataframe_to_rows(idle_df, index=False, header=True):
        ws_idle.append(r)
    _adjust_column_widths(ws_idle)
    
    if pie_idle_png:
        img1 = XLImage(io.BytesIO(pie_idle_png))
        img1.width = 600
        img1.height = 400
        ws_idle.add_image(img1, 'F2')

    # время выпуска
    ws_comp = wb.create_sheet("Время выпуска")
    comp_df = completion_by_job.rename(columns={
        "Job": "Деталь",
        "Finish": "Время завершения"
    })
    for r in dataframe_to_rows(comp_df, index=False, header=True):
        ws_comp.append(r)
    _adjust_column_widths(ws_comp)
    
    if pie_comp_png:
        img2 = XLImage(io.BytesIO(pie_comp_png))
        img2.width = 600
        img2.height = 400
        ws_comp.add_image(img2, 'F2')

    # ганта до
    ws_gantt_before = wb.create_sheet("Ганта_до")
    gantt_before_ru = gantt_before.rename(columns={
        "Machine": "Станок",
        "Job": "Деталь",
        "Start": "Начало",
        "Finish": "Окончание"
    })
    for r in dataframe_to_rows(gantt_before_ru, index=False, header=True):
        ws_gantt_before.append(r)
    _adjust_column_widths(ws_gantt_before)

    # ганта после
    ws_gantt_after = wb.create_sheet("Ганта_после")
    gantt_after_ru = gantt_after.rename(columns={
        "Machine": "Станок",
        "Job": "Деталь",
        "Start": "Начало",
        "Finish": "Окончание"
    })
    for r in dataframe_to_rows(gantt_after_ru, index=False, header=True):
        ws_gantt_after.append(r)
    _adjust_column_widths(ws_gantt_after)

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _adjust_column_widths(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = min(adjusted_width, 30)